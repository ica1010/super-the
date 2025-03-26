from personnels.models import User
from pos.models  import Caisse, Category, Order, Product
# Create your views here.
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Sum
from crm.models import Client, Fournisseur
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

@login_required(login_url='login')
def Dashboard(request):
    categories = Category.objects.all()
    products = Product.objects.all()
    clients = Client.objects.filter(disponible=True)
    today = timezone.now().date() 
    try:
        caisse = Caisse.objects.get(user=request.user)
    except :
        caisse = {}
    # Commandes complètes
    completed_orders = Order.objects.filter(status='Validé',status_de_paiement="non soldée").prefetch_related('items__product')
    # Commandes payées
    paused_orders = Order.objects.filter(status = 'Mis en pause')
    
    
    all_orders = completed_orders.union(Order.objects.filter(status_de_paiement='soldée').prefetch_related('items__product')) 
    
    
    completed_orders_total = Order.objects.filter(status='Validé').annotate(
        order_total=Sum('items__product__price', field='items__quantity')
    ).aggregate(Sum('order_total'))['order_total__sum'] or 0
    
    
    close_orders_total =  Order.objects.filter(status_de_payement='soldée').annotate(
        order_total=Sum('items__product__price', field='items__quantity')
    ).aggregate(Sum('order_total'))['order_total__sum'] or 0
    
    
    context = {
        'categories': categories,
        'clients': clients,
        'paused_orders': paused_orders,
        'completed_orders':completed_orders,
        'all_orders':all_orders,
        'completed_orders_total':completed_orders_total,
        'close_orders_total':close_orders_total,
        'products':products,
        'caisse':caisse,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'partials/dashboard.html', context)
        
    return render(request, 'pages/dashboard.html',context)

    

@login_required(login_url='login')
def VendorDashboard(request):
    clients = Client.objects.filter(disponible=True)
    categories = Category.objects.all()
    products = Product.objects.all()
    today = timezone.now().date() 
      # Commandes complètes
    completed_orders = Order.objects.filter(status='Validé', created_at__date=today).prefetch_related('items__product')
    # Commandes payées
    paused_orders = Order.objects.filter(status = 'Mis en pause')
    
    close_orders = Order.objects.filter(status_de_payement='soldée', created_at__date=today).prefetch_related('items__product')
    all_orders = completed_orders.union(close_orders) 
    completed_orders_total = Order.objects.filter(status='Validé', created_at__date=today).annotate(
        order_total=Sum('items__product__price', field='items__quantity')
    ).aggregate(Sum('order_total'))['order_total__sum'] or 0
    close_orders_total =  Order.objects.filter(status_de_payement='soldée').annotate(
        order_total=Sum('items__product__price', field='items__quantity')
    ).aggregate(Sum('order_total'))['order_total__sum'] or 0
    
    
    context = {
        'clients': clients,
        'categories': categories,
        'paused_orders': paused_orders,
        'completed_orders':completed_orders,
        'close_orders':close_orders,
        'all_orders':all_orders,
        'completed_orders_total':completed_orders_total,
        'close_orders_total':close_orders_total,
        'products':products,
    }


    return render(request, 'pages/vendor-dashbord.html',context)
 
 
def Achat(request):
    fournisseurs = Fournisseur.objects.filter(disponible=True)
    categories = Category.objects.all()
    products = Product.objects.filter(product_type = True)
    today = timezone.now().date() 
    
    
    context = {
        'fournisseurs': fournisseurs,
        'categories': categories,
        'products':products,
    }

    
    return render(request, 'pages/achat.html', context)
       
        
def get_order_data(request):
    # Agrégation des totaux par statut
    completed_orders_total = Order.objects.filter(status='Validé').annotate(
        order_total=Sum('items__product__price', field='items__quantity')
    ).aggregate(Sum('order_total'))['order_total__sum'] or 0

    # Calculer le total des commandes payées
    close_orders_total = Order.objects.filter(status_de_payement='soldée').annotate(
        order_total=Sum('items__product__price', field='items__quantity')
    ).aggregate(Sum('order_total'))['order_total__sum'] or 0
    
    data = {
        'labels': ['Commandes complétées', 'Commandes payées'],
        'totals': [completed_orders_total, close_orders_total]
    }
    return JsonResponse(data)

    
def get_total_orders_by_status(request, status):
    today = timezone.now().date()
    orders_today = Order.objects.filter(status=status, created_at__date=today)
    
    total = orders_today.annotate(order_total=Sum('items__product__price', field='items__quantity')).aggregate(Sum('order_total'))['order_total__sum'] or 0
    
    # Formater le total avec des milliers séparés et deux décimales
    formatted_total = f"{total:,.2f}"
    
    # Retourner un JSON avec la valeur du total
    return JsonResponse({
        'total': total,
        'formatted_total': formatted_total
    })
    
    

def Cuisinier(request):
    
    orders = Order.objects.filter(status_de_paiement = 'non soldée').prefetch_related('items__product')
    
    serveurs = User.objects.filter(commandes__status_de_paiement='non soldée').distinct().prefetch_related('commandes')

    # Vérifie si la requête provient de HTMX
    if request.headers.get('HX-Request'):
        return render(request, 'partials/cuisinier.html', {'orders': orders,'serveurs': serveurs})

    return render(request, 'pages/cuisinier.html',  {'orders': orders,'serveurs': serveurs})



def payement(request, order_id):
    current_order = get_object_or_404(Order, id=order_id)
    url = request.build_absolute_uri() 
    if request.method == "POST":
        # current_order.status_de_payement = 'soldée'
        current_order.status_de_paiement = 'soldée'
        current_order.save()
        try:
            # Mettre à jour le statut de la commande actuelle à 'PAYE'
            
            
            # Récupérer toutes les commandes complétées à l'exception de la commande actuelle
            orders = Order.objects.filter(status_de_paiement='soldée').exclude(id=order_id).prefetch_related('items__product')

            # Vérifier si la requête est une requête HTMX pour un rendu partiel
            if request.headers.get('HX-Request'):
                return redirect('cuisinier')

            # Rediriger vers une autre page en cas de requête standard
            return redirect('cuisinier')  # remplace 'order_success' par ta vue ou URL de redirection

        except Exception as e:
            # Gérer les exceptions et retourner une réponse JSON en cas de requête HTMX
            if request.headers.get('HX-Request'):
                return JsonResponse({'error': str(e)}, status=400)
            return HttpResponse(e)

    # Retourner une réponse 405 si la méthode n'est pserver=request.userserver=request.useras POST
    return redirect("cuisinier")


def export_table_to_excel(request):
    # Create a workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Table Data"

    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Set headers
    headers = ["Column 1", "Column 2"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add data rows
    table_data = [...]  # Replace with your actual data
    for row_num, row_data in enumerate(table_data, 2):
        worksheet.cell(row=row_num, column=1, value=row_data['column1'])
        worksheet.cell(row=row_num, column=2, value=row_data['column2'])

    # Set response for Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="table.xlsx"'